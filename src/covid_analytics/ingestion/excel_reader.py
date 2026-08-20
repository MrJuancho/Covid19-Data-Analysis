"""Lectura acotada de Excel: detección de encabezado por texto + límite de filas
(FR-001), evitando el rango fantasma documentado en docs/audit_legacy.md.

Los encabezados reales del archivo fuente no son necesariamente una sola fila:
`SEGUIMIENTO DE CASOS COVID 19` tiene 1 fila con sufijos descriptivos
verbosos (ej. "SEXO (M/F)"), mientras que ` NOMINAL DE HOSPITALIZADOS` tiene
encabezados combinados en hasta 3 filas con celdas agrupadas (ej. "EDAD" en
una fila y "F"/"M" en la siguiente). `filas_encabezado` deja que el llamador
declare cuántas filas debe combinar esta función para ese caso.

Todo el módulo usa exclusivamente `openpyxl` en modo `read_only=True`: la API
de `pandas.read_excel(engine="openpyxl")` invoca `load_workbook` SIN modo
`read_only`, lo que materializa en memoria las celdas (con estilos) de las
~16,000 columnas fantasma de TODAS las hojas del libro por cada llamada — con
el archivo real esto disparó picos de memoria de varios GB muy por encima del
límite de SC-002 (500 MB). `read_only=True` transmite fila por fila sin ese
costo.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

_MAX_FILAS_BUSQUEDA_HEADER = 50
# Acota el ancho al buscar/combinar encabezados: el "rango usado" de Excel
# puede reportar miles de columnas fantasma (docs/audit_legacy.md); sin este
# límite, el forward-fill de _combinar_encabezado arrastraría la última
# etiqueta real indefinidamente hacia esas columnas fantasma, generando
# nombres de columna duplicados.
_MAX_COLUMNAS_ENCABEZADO = 200


class EncabezadoNoEncontradoError(RuntimeError):
    """No se localizó una fila de encabezado que contenga las columnas clave."""


def _leer_filas(
    ruta: Path, sheet_name: str, min_row: int, max_row: int
) -> list[tuple[object, ...]]:
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return list(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                max_col=_MAX_COLUMNAS_ENCABEZADO,
                values_only=True,
            )
        )
    finally:
        wb.close()


def _celda_coincide(texto_celda: str, clave: str) -> bool:
    """Compara con conciencia de límite de palabra: `clave` debe ser el
    contenido completo de la celda o un prefijo seguido de un carácter no
    alfanumérico (espacio, paréntesis, coma...). Evita falsos positivos como
    "HOSPITAL" dentro de "...PACIENTES HOSPITALIZADOS COMO..." (título de la
    hoja Nominal)."""
    t = texto_celda.strip().upper()
    if t == clave:
        return True
    if t.startswith(clave):
        resto = t[len(clave) :]
        return not resto or not resto[0].isalnum()
    return False


def _fila_coincide(fila: tuple[object, ...], clave: str) -> bool:
    return any(v is not None and _celda_coincide(str(v), clave) for v in fila)


def _encontrar_fila_encabezado(
    ruta: Path, sheet_name: str, columnas_clave: list[str], filas_encabezado: int
) -> int:
    """Devuelve el índice 0-based de la primera fila del bloque de encabezado.

    Ancla la búsqueda exigiendo que la PRIMERA `columna_clave` aparezca (por
    celda completa, con límite de palabra) en la primera fila de la ventana
    candidata; la ventana completa de `filas_encabezado` filas debe contener
    todas las `columnas_clave`, cada una en alguna celda de esa ventana."""
    claves_norm = [c.upper() for c in columnas_clave]
    ancla = claves_norm[0]
    filas = _leer_filas(ruta, sheet_name, 1, _MAX_FILAS_BUSQUEDA_HEADER)
    for inicio in range(len(filas) - filas_encabezado + 1):
        if not _fila_coincide(filas[inicio], ancla):
            continue
        ventana = filas[inicio : inicio + filas_encabezado]
        if all(any(_fila_coincide(fila, clave) for fila in ventana) for clave in claves_norm):
            return inicio
    raise EncabezadoNoEncontradoError(
        f"No se encontró bloque de encabezado ({filas_encabezado} fila(s)) con "
        f"{columnas_clave} en '{sheet_name}' dentro de las primeras "
        f"{_MAX_FILAS_BUSQUEDA_HEADER} filas."
    )


def _combinar_encabezado(ventana: list[tuple[object, ...]]) -> list[str]:
    """Combina una ventana de filas de encabezado en una lista de nombres de
    columna planos, ej. `["EDAD", None]` + `["F", "M"]` -> `["EDAD | F",
    "EDAD | M"]`, replicando cómo se ve una celda combinada de Excel al
    desagruparla (forward-fill hacia la derecha dentro de cada fila).

    El forward-fill de un nivel se reinicia en cualquier columna donde un
    nivel SUPERIOR (una fila más arriba en la ventana) tiene su propio valor
    nuevo — si no, la etiqueta de un grupo lateral (ej. "M" de EDAD|M)
    se filtraría hacia columnas de otro grupo sin relación (ej. "LABORATORIO
    | RESULTADO" terminaría heredando "M")."""
    n_cols = max((len(fila) for fila in ventana), default=0)
    crudos = [[(fila[c] if c < len(fila) else None) for c in range(n_cols)] for fila in ventana]

    limite_nivel_superior = [False] * n_cols
    ffilled: list[list[str | None]] = []
    for crudos_nivel in crudos:
        fila_ffill: list[str | None] = []
        carry: str | None = None
        for col in range(n_cols):
            valor_crudo = crudos_nivel[col]
            if valor_crudo is not None:
                carry = str(valor_crudo).strip()
                fila_ffill.append(carry)
                limite_nivel_superior[col] = True
            elif limite_nivel_superior[col]:
                carry = None
                fila_ffill.append(None)
            else:
                fila_ffill.append(carry)
        ffilled.append(fila_ffill)

    # Recorta al ancho real: cualquier columna después de la última con un
    # valor propio (no heredado por ffill) en algún nivel es una columna
    # fantasma de Excel — conservarla produciría nombres duplicados.
    columnas_con_valor_propio = [c for c in range(n_cols) if limite_nivel_superior[c]]
    ancho_real = (columnas_con_valor_propio[-1] + 1) if columnas_con_valor_propio else 0

    nombres: list[str] = []
    for col in range(ancho_real):
        valores_col = (f[col] for f in ffilled)
        partes = [v for v in valores_col if v is not None]
        nombres.append(" | ".join(dict.fromkeys(partes)) or f"__col_{col}__")
    return nombres


_MAX_SALTO_FILAS_EN_BLANCO = 10


def _localizar_filas_de_datos(
    ruta: Path, sheet_name: str, desde_1based: int, columna_clave_idx: int
) -> tuple[int, int]:
    """Devuelve (primera_fila_datos_1based, num_filas): salta filas donde la
    columna clave está vacía (fila espaciadora, o una sub-fila de encabezado
    que sólo tiene contenido en otra columna — observado en la hoja
    Seguimiento real, ver docs/audit_legacy.md) y luego cuenta filas
    contiguas con datos en esa columna, deteniéndose en la primera fila
    vacía para no arrastrar el rango fantasma de Excel."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        it = ws.iter_rows(min_row=desde_1based, values_only=True)
        fila_actual = desde_1based
        saltos = 0
        fila = next(it, None)
        while fila is not None and saltos < _MAX_SALTO_FILAS_EN_BLANCO:
            valor = fila[columna_clave_idx] if columna_clave_idx < len(fila) else None
            if valor is not None:
                break
            saltos += 1
            fila_actual += 1
            fila = next(it, None)

        primera_fila_datos = fila_actual
        conteo = 0
        while fila is not None:
            valor = fila[columna_clave_idx] if columna_clave_idx < len(fila) else None
            if valor is None:
                break
            conteo += 1
            fila = next(it, None)
        return primera_fila_datos, conteo
    finally:
        wb.close()


def _leer_datos(
    ruta: Path, sheet_name: str, primera_fila_1based: int, num_filas: int, ancho: int
) -> list[tuple[object, ...]]:
    """Lee exactamente `num_filas` x `ancho` celdas de datos, en modo
    `read_only`, sin materializar columnas fantasma ni objetos de estilo."""
    if num_filas == 0:
        return []
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        filas = ws.iter_rows(
            min_row=primera_fila_1based,
            max_row=primera_fila_1based + num_filas - 1,
            max_col=ancho,
            values_only=True,
        )
        return list(filas)
    finally:
        wb.close()


def leer_hoja_acotada(
    excel_path: str | Path,
    sheet_name: str,
    columnas_clave: list[str],
    filas_encabezado: int = 1,
) -> pd.DataFrame:
    """Lee una hoja localizando el bloque de encabezado real por búsqueda de
    texto (no `skiprows` fijo) y acotando filas/columnas a lo estrictamente
    detectado, para no materializar el rango fantasma de Excel. Usa
    `openpyxl` en modo `read_only` de principio a fin (ver docstring del
    módulo) — nunca delega a `pandas.read_excel(engine="openpyxl")`.
    `filas_encabezado` combina encabezados repartidos en varias filas (ej.
    hoja Nominal)."""
    ruta = Path(excel_path)
    inicio_0based = _encontrar_fila_encabezado(ruta, sheet_name, columnas_clave, filas_encabezado)
    ventana_encabezado = _leer_filas(
        ruta, sheet_name, inicio_0based + 1, inicio_0based + filas_encabezado
    )
    nombres_columna = _combinar_encabezado(ventana_encabezado)

    columna_clave_idx = next(
        i for i, n in enumerate(nombres_columna) if _celda_coincide(n, columnas_clave[0].upper())
    )

    desde_1based = inicio_0based + filas_encabezado + 1
    primera_fila_datos_1based, num_filas = _localizar_filas_de_datos(
        ruta, sheet_name, desde_1based, columna_clave_idx
    )

    datos = _leer_datos(
        ruta, sheet_name, primera_fila_datos_1based, num_filas, len(nombres_columna)
    )
    return pd.DataFrame(datos, columns=nombres_columna)
