"""Generador de un .xlsx 100% sintético que replica el layout de las 3 hojas
reales documentado en docs/audit_legacy.md, sin ningún dato real de pacientes.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

# Fila de encabezado real dentro de "SEGUIMIENTO DE CASOS COVID 19" (0-based),
# igual que en docs/audit_legacy.md (el legacy la descarta por accidente).
SEGUIMIENTO_HEADER_ROW_0BASED = 12

SEGUIMIENTO_COLUMNS = [
    "NO. (consecutivo)",
    "CASO (nombre,apellido paterno, apellido materno)",
    "SEXO (M/F)",
    "EDAD (ej: anos: 3;meses-3M)",
    "MUNICIPIO/PAIS RESIDENCIA",
    "UNIDAD NOTIFICANTE",
    "JURISDICCION",
    "DERECHOHABIENCIA (IMSS, ISSSTE, ISSEMYM, SEDENA, NINGUNO)",
    "FECHA DE NOTIFICACION (dd/mm/aaaa)",
    "SIGNOS Y SINTOMAS",
    "FECHA TOMA DE MUESTRA (dd/mm/aaaa)",
    "LABORATORIO",
    "RESULTADO (POSITIVO, NEGATIVO, PENDIENTE)",
    "FECHA DE RESULTADO  (dd/mm/aaaa)",
    "ESTATUS",
    "PAIS DE PROCEDENCIA",
]

# La hoja Nominal real tiene un encabezado combinado en 3 filas: la fila 0
# (relativa) trae las etiquetas de grupo/columnas simples (HOSPITAL, NO.,
# NOMBRE DEL PACIENTE, EDAD, LABORATORIO), la fila 1 las sub-columnas de EDAD
# (F/M), y la fila 2 el bloque de laboratorio (RESULTADO, FECHA DE
# RESULTADO), como celdas combinadas. Ver docs/audit_legacy.md.
NOMINAL_HEADER_ROW_0BASED = 5
NOMINAL_FILAS_ENCABEZADO = 3

# Celdas de encabezado a escribir: (fila_relativa, columna_0based, texto).
# Incluye las etiquetas de grupo ("LABORATORIO") que en el archivo real
# delimitan dónde termina el "forward-fill" de la celda combinada anterior.
_NOMINAL_CELDAS_ENCABEZADO: list[tuple[int, int, str]] = [
    (0, 0, "HOSPITAL"),
    (0, 1, "NO."),
    (0, 2, "NOMBRE DEL PACIENTE"),
    (0, 3, "EDAD"),
    (1, 3, "F"),
    (1, 4, "M"),
    (0, 5, "FECHA DE INGRESO O DE ATENCION"),
    (0, 6, "LABORATORIO"),
    (2, 6, "RESULTADO"),
    (2, 7, "FECHA DE RESULTADO"),
]
# Claves "hoja" (leaf) usadas por los llamadores en las filas de datos ->
# columna 0-based donde escribir el valor.
_NOMINAL_CLAVE_A_COLUMNA: dict[str, int] = {
    "HOSPITAL": 0,
    "NO.": 1,
    "NOMBRE DEL PACIENTE": 2,
    "F": 3,
    "M": 4,
    "FECHA DE INGRESO O DE ATENCION": 5,
    "RESULTADO": 6,
    "FECHA DE RESULTADO": 7,
}
NOMINAL_COLUMNS = list(_NOMINAL_CLAVE_A_COLUMNA.keys())

RED_NEGATIVA_COLUMNS = [
    "FECHA 1",
    "HOSPITAL",
    "INSTITUCION",
    "NUEVOS CASOS",
    "CASOS ACUMULADOS",
]


def _escribir_encabezado(ws, fila_1based: int, columnas: list[str]) -> None:
    for idx, nombre in enumerate(columnas, start=1):
        ws.cell(row=fila_1based, column=idx, value=nombre)


def construir_excel_sintetico(
    destino: Path,
    filas_seguimiento: list[dict[str, object]],
    filas_nominal: list[dict[str, object]],
    filas_red_negativa: list[dict[str, object]] | None = None,
    simular_rango_fantasma: bool = False,
) -> Path:
    """Crea un .xlsx sintético con las 3 hojas auditadas.

    `filas_*` son listas de dicts columna->valor (valores 100% sintéticos,
    nunca datos reales de pacientes). Las claves de `filas_seguimiento` son
    nombres de `SEGUIMIENTO_COLUMNS`; las de `filas_nominal` son las claves
    "hoja" de `NOMINAL_COLUMNS` (ej. "F"/"M" para EDAD, "RESULTADO", etc.).
    `simular_rango_fantasma` toca una celda lejana para inflar el rango usado
    del sheet, replicando el bug de Excel documentado en docs/audit_legacy.md.
    """
    wb = Workbook()

    ws_red = wb.active
    ws_red.title = "RED NEGATIVA"
    _escribir_encabezado(ws_red, 1, RED_NEGATIVA_COLUMNS)
    for fila_idx, fila in enumerate(filas_red_negativa or [], start=2):
        for col_idx, nombre in enumerate(RED_NEGATIVA_COLUMNS, start=1):
            ws_red.cell(row=fila_idx, column=col_idx, value=fila.get(nombre))

    ws_seg = wb.create_sheet("SEGUIMIENTO DE CASOS COVID 19")
    header_row_1based = SEGUIMIENTO_HEADER_ROW_0BASED + 1
    _escribir_encabezado(ws_seg, header_row_1based, SEGUIMIENTO_COLUMNS)
    for offset, fila in enumerate(filas_seguimiento):
        fila_idx = header_row_1based + 1 + offset
        for col_idx, nombre in enumerate(SEGUIMIENTO_COLUMNS, start=1):
            ws_seg.cell(row=fila_idx, column=col_idx, value=fila.get(nombre))
    if simular_rango_fantasma:
        ws_seg.cell(row=100_000, column=len(SEGUIMIENTO_COLUMNS) + 5, value=" ")

    ws_nom = wb.create_sheet(" NOMINAL DE HOSPITALIZADOS")
    for fila_rel, col_0based, texto in _NOMINAL_CELDAS_ENCABEZADO:
        ws_nom.cell(
            row=NOMINAL_HEADER_ROW_0BASED + fila_rel + 1,
            column=col_0based + 1,
            value=texto,
        )
    primera_fila_datos_1based = NOMINAL_HEADER_ROW_0BASED + NOMINAL_FILAS_ENCABEZADO + 1
    for offset, fila in enumerate(filas_nominal):
        fila_idx = primera_fila_datos_1based + offset
        for clave, col_0based in _NOMINAL_CLAVE_A_COLUMNA.items():
            ws_nom.cell(row=fila_idx, column=col_0based + 1, value=fila.get(clave))
    if simular_rango_fantasma:
        ws_nom.cell(row=100_000, column=len(_NOMINAL_CLAVE_A_COLUMNA) + 5, value=" ")

    wb.save(destino)
    return destino
